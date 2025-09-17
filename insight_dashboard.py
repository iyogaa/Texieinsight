import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
from io import BytesIO

def process_merged_cells(df, sheet_name, excel_file):
    """
    Process merged cells in Excel sheets by expanding merged values to all cells in the merge range
    """
    try:
        # Load the workbook with openpyxl to access merged cell information
        wb = load_workbook(filename=BytesIO(excel_file.read()), data_only=True)
        sheet = wb[sheet_name]
        
        # Get all merged cell ranges
        merged_ranges = sheet.merged_cells.ranges
        
        # Create a copy of the dataframe to modify
        processed_df = df.copy()
        
        # Process each merged range
        for merged_range in merged_ranges:
            # Get the top-left cell of the merged range (this contains the value)
            top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
            value = top_left_cell.value
            
            # Apply the value to all cells in the merged range
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    # Convert Excel row/col to dataframe index/column
                    df_row = row - 2  # Subtract 2 because Excel rows start at 1 and header is row 1
                    df_col = col - 1  # Subtract 1 because Excel columns start at 1
                    
                    # Only update if within dataframe bounds
                    if 0 <= df_row < len(processed_df) and 0 <= df_col < len(processed_df.columns):
                        col_name = processed_df.columns[df_col]
                        processed_df.iloc[df_row, df_col] = value
        
        return processed_df
    except Exception as e:
        st.warning(f"Could not process merged cells for {sheet_name}: {str(e)}")
        return df

def insight_dashboard_app():
    st.markdown("""
        <style>
        .stApp {
            background-color: #0d1117;
            color: #c9d1d9;
            font-family: 'Segoe UI', sans-serif;
        }
        h1, h2, h3 {
            color: #f0f6fc;
        }
        .stSidebar {
            background-color: #161b22;
        }
        .stDataFrame {
            background-color: #161b22;
            color: #c9d1d9;
        }
        .stButton>button, .stDownloadButton>button {
            background-color: #238636;
            color: #ffffff;
            border: none;
            padding: 0.5em 1em;
            border-radius: 6px;
            font-weight: bold;
        }
        .stButton>button:hover, .stDownloadButton>button:hover {
            background-color: #2ea043;
            color: #ffffff;
        }
        .stMetric {
            background-color: #161b22;
            padding: 12px;
            border-radius: 8px;
            color: #f0f6fc;
            border: 1px solid #30363d;
            box-shadow: 0 0 5px #238636;
        }
        .plot-container {
            background-color: #161b22;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 20px;
        }
        .highlight {
            background-color: #238636;
            color: white;
            padding: 5px;
            border-radius: 5px;
        }
        </style>
    """, unsafe_allow_html=True)

    # Sidebar elements
    st.sidebar.markdown("### 📝 Submit a New Task")
    st.sidebar.markdown("[Open Google Form](https://forms.gle/your-form-id)")

    st.sidebar.markdown("### 📤 Upload Excel File")
    uploaded_file = st.sidebar.file_uploader("Choose an Excel file", type=["xlsx"])

    # Client tabs
    client_tabs = ["HDVI_70", "AllTrans_87", "RISCOM_88", "Method", "Foresight_127"]
    
    # Define the logic for each client tab
    client_config = {
        "HDVI_70": {
            "category_column": "Class Name",
            "count_column": "Class Name"
        },
        "AllTrans_87": {
            "category_column": "Class Name",
            "count_column": "Class Name"
        },
        "RISCOM_88": {
            "category_column": "Class Name",
            "count_column": "Class Name"
        },
        "Method": {
            "category_column": "Category",
            "count_column": "Category"
        },
        "Foresight_127": {
            "category_column": "Category",
            "count_column": "Category"
        }
    }

    if uploaded_file:
        try:
            # Store the file content for merged cell processing
            excel_content = uploaded_file.read()
            uploaded_file.seek(0)  # Reset file pointer for pandas
            
            # Read all sheets
            dfs = {}
            for client in client_tabs:
                try:
                    df_sheet = pd.read_excel(uploaded_file, sheet_name=client)
                    # Process merged cells for this sheet
                    df_sheet = process_merged_cells(df_sheet, client, BytesIO(excel_content))
                    dfs[client] = df_sheet
                except Exception as e:
                    st.sidebar.warning(f"Could not load sheet {client}: {str(e)}")
                    continue

            # Client selection
            st.sidebar.title("📁 Select Client")
            available_clients = [c for c in client_tabs if c in dfs]
            if not available_clients:
                st.error("No valid client sheets found in the uploaded file.")
                return
                
            selected_client = st.sidebar.selectbox("Choose a client tab", available_clients)
            df = dfs[selected_client].copy()
            
            # Get configuration for selected client
            config = client_config[selected_client]
            category_column = config["category_column"]
            count_column = config["count_column"]

            # Handle any remaining NaN values in key columns
            key_columns = ['Request Id', 'Verify_user', 'Upload_Status', category_column]
            for col in key_columns:
                if col in df.columns:
                    # Forward fill NaN values
                    df[col] = df[col].fillna(method='ffill')
            
            # Date processing with error handling
            try:
                df['Creation Date'] = pd.to_datetime(df['Creation Date'], errors='coerce')
            except Exception as e:
                st.warning(f"Date processing issue: {e}. Creating placeholder dates.")
                df['Creation Date'] = pd.date_range(start='2023-01-01', periods=len(df), freq='D')
            
            # Handle time column - check if it exists and process accordingly
            if 'Creation Time' in df.columns:
                try:
                    df['Creation Time'] = pd.to_datetime(df['Creation Time'], errors='coerce').dt.time
                except:
                    df['Creation Time'] = pd.NaT
            else:
                df['Creation Time'] = pd.NaT
            
            # Set default date range (last 7 days)
            default_end = datetime.today()
            default_start = default_end - timedelta(days=7)
            
            st.sidebar.markdown("### 📅 Select Date Range")
            start_date = st.sidebar.date_input("Start Date", value=default_start)
            end_date = st.sidebar.date_input("End Date", value=default_end)

            # Filter options based on category column
            if category_column in df.columns:
                # Remove NaN values and get unique categories
                class_options = [x for x in df[category_column].dropna().unique() if str(x) != 'nan']
                selected_class = st.sidebar.selectbox(f"🎓 Filter by {category_column}", ["All"] + class_options)
            else:
                selected_class = "All"
                st.sidebar.warning(f"Column '{category_column}' not found in data")

            # Status filter
            if 'Upload_Status' in df.columns:
                status_options = [x for x in df['Upload_Status'].dropna().unique() if str(x) != 'nan']
                selected_status = st.sidebar.multiselect(
                    "Filter by Status", 
                    options=status_options,
                    default=status_options
                )
            else:
                selected_status = []

            # Apply filters
            filtered_df = df[(df['Creation Date'] >= pd.to_datetime(start_date)) &
                            (df['Creation Date'] <= pd.to_datetime(end_date))]
            
            if selected_class != "All":
                filtered_df = filtered_df[filtered_df[category_column] == selected_class]
            
            if selected_status:
                filtered_df = filtered_df[filtered_df['Upload_Status'].isin(selected_status)]

            # Display dashboard title
            st.title(f"📊 {selected_client} Dashboard")
            st.subheader(f"🗓️ Results for {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}")

            # Key metrics - with careful counting
            col1, col2, col3, col4 = st.columns(4)
            
            # Count unique request IDs carefully
            if 'Request Id' in filtered_df.columns:
                # Remove NaN values before counting
                request_ids = filtered_df['Request Id'].dropna()
                unique_tasks = request_ids.nunique()
            else:
                unique_tasks = 0
                
            col1.metric("🧮 Unique Tasks", unique_tasks)
            
            # Count successful uploads
            if 'Upload_Status' in filtered_df.columns:
                success_count = (filtered_df['Upload_Status'] == 'Success').sum()
            else:
                success_count = 0
            col2.metric("✅ Submitted", success_count)
            
            # Count failed uploads
            if 'Upload_Status' in filtered_df.columns:
                failed_count = (filtered_df['Upload_Status'] == 'Failed').sum()
            else:
                failed_count = 0
            col3.metric("❌ Failed", failed_count)
            
            # Count unique verifiers
            if 'Verify_user' in filtered_df.columns:
                verifiers = filtered_df['Verify_user'].dropna()
                unique_verifiers = verifiers.nunique()
            else:
                unique_verifiers = 0
            col4.metric("👥 Verifiers", unique_verifiers)

            # Create tabs for different views
            tab1, tab2, tab3 = st.tabs(["Client & Class Summary", "Detailed View", "Data Quality"])

            with tab1:
                st.subheader("Client & Class Summary")
                
                # Client-wise summary
                client_summary_data = []
                for client in client_tabs:
                    if client in dfs:
                        client_df = dfs[client].copy()
                        
                        # Handle any remaining NaN values
                        client_config_data = client_config[client]
                        client_category_col = client_config_data["category_column"]
                        
                        key_columns = ['Request Id', 'Verify_user', 'Upload_Status', client_category_col]
                        for col in key_columns:
                            if col in client_df.columns:
                                client_df[col] = client_df[col].fillna(method='ffill')
                        
                        # Date processing with error handling
                        try:
                            client_df['Creation Date'] = pd.to_datetime(client_df['Creation Date'], errors='coerce')
                        except:
                            client_df['Creation Date'] = pd.date_range(start='2023-01-01', periods=len(client_df), freq='D')
                        
                        # Filter by date
                        client_filtered = client_df[(client_df['Creation Date'] >= pd.to_datetime(start_date)) &
                                                  (client_df['Creation Date'] <= pd.to_datetime(end_date))]
                        
                        # Count unique requests carefully
                        if 'Request Id' in client_filtered.columns:
                            client_request_ids = client_filtered['Request Id'].dropna()
                            unique_requests = client_request_ids.nunique()
                        else:
                            unique_requests = 0
                            
                        # Count successes carefully
                        if 'Upload_Status' in client_filtered.columns:
                            success_count = (client_filtered['Upload_Status'] == 'Success').sum()
                        else:
                            success_count = 0
                        
                        client_summary_data.append({
                            'Client': client,
                            'Unique Tasks': unique_requests,
                            'Success Count': success_count,
                            'Success Rate': (success_count / unique_requests * 100) if unique_requests > 0 else 0
                        })
                
                client_summary_df = pd.DataFrame(client_summary_data)
                
                if not client_summary_df.empty:
                    # Display client summary
                    st.markdown("#### 📋 Client-wise Summary")
                    fig = px.bar(
                        client_summary_df, 
                        x='Client', 
                        y='Unique Tasks',
                        color='Success Rate',
                        color_continuous_scale='Viridis',
                        title='Tasks by Client'
                    )
                    fig.update_layout({
                        'plot_bgcolor': 'rgba(0, 0, 0, 0)',
                        'paper_bgcolor': 'rgba(0, 0, 0, 0)',
                        'font': {'color': '#c9d1d9'},
                        'xaxis': {'gridcolor': '#30363d'},
                        'yaxis': {'gridcolor': '#30363d'}
                    })
                    st.plotly_chart(fig, use_container_width=True)
                    
                    # Display client summary table
                    st.dataframe(client_summary_df)
                
                # Class/Category-wise summary for selected client
                if category_column in filtered_df.columns:
                    st.markdown(f"#### 🎓 {category_column}-wise Summary for {selected_client}")
                    
                    # Group by category and count carefully
                    class_summary = filtered_df.groupby(category_column).agg({
                        'Request Id': lambda x: x.dropna().nunique(),
                        'Upload_Status': lambda x: (x == 'Success').sum()
                    }).reset_index()
                    
                    class_summary.columns = [category_column, 'Unique Tasks', 'Success Count']
                    class_summary['Success Rate'] = (class_summary['Success Count'] / class_summary['Unique Tasks'] * 100).round(2)
                    class_summary = class_summary.sort_values('Unique Tasks', ascending=False)
                    
                    # Display class summary chart
                    fig = px.bar(
                        class_summary, 
                        x=category_column, 
                        y='Unique Tasks',
                        color='Success Rate',
                        color_continuous_scale='Viridis',
                        title=f'{category_column}-wise Task Distribution'
                    )
                    fig.update_layout({
                        'plot_bgcolor': 'rgba(0, 0, 0, 0)',
                        'paper_bgcolor': 'rgba(0, 0, 0, 0)',
                        'font': {'color': '#c9d1d9'},
                        'xaxis': {'gridcolor': '#30363d', 'tickangle': 45},
                        'yaxis': {'gridcolor': '#30363d'}
                    })
                    st.plotly_chart(fig, use_container_width=True)
                    
                    # Display class summary table
                    st.dataframe(class_summary)

            with tab2:
                st.subheader("Detailed Task View")
                
                # Upload Status Breakdown
                if 'Upload_Status' in filtered_df.columns:
                    status_counts = filtered_df['Upload_Status'].value_counts().reset_index()
                    status_counts.columns = ['Status', 'Count']
                    
                    # Create a pie chart for status distribution
                    fig = px.pie(
                        status_counts, 
                        values='Count', 
                        names='Status', 
                        title='Upload Status Distribution',
                        color_discrete_sequence=px.colors.qualitative.Set3
                    )
                    fig.update_layout({
                        'plot_bgcolor': 'rgba(0, 0, 0, 0)',
                        'paper_bgcolor': 'rgba(0, 0, 0, 0)',
                        'font': {'color': '#c9d1d9'}
                    })
                    st.plotly_chart(fig, use_container_width=True)
                
                # Daily Task Trend
                daily_trend = filtered_df.groupby('Creation Date')['Request Id'].count().reset_index()
                daily_trend.columns = ['Date', 'Task Count']
                
                fig = px.line(
                    daily_trend, 
                    x='Date', 
                    y='Task Count', 
                    title='Daily Task Trend',
                    markers=True
                )
                fig.update_layout({
                    'plot_bgcolor': 'rgba(0, 0, 0, 0)',
                    'paper_bgcolor': 'rgba(0, 0, 0, 0)',
                    'font': {'color': '#c9d1d9'},
                    'xaxis': {'gridcolor': '#30363d'},
                    'yaxis': {'gridcolor': '#30363d'}
                })
                st.plotly_chart(fig, use_container_width=True)
                
                # Verifier performance
                if 'Verify_user' in filtered_df.columns:
                    verifier_summary = filtered_df['Verify_user'].value_counts().reset_index()
                    verifier_summary.columns = ['Verifier', 'Tasks Verified']
                    verifier_summary = verifier_summary.sort_values(by='Tasks Verified', ascending=False)
                    
                    fig = px.bar(
                        verifier_summary.head(10),  # Show top 10 only
                        x='Verifier', 
                        y='Tasks Verified', 
                        title="Top 10 Verifiers",
                        color='Tasks Verified',
                        color_continuous_scale='Plasma'
                    )
                    fig.update_layout({
                        'plot_bgcolor': 'rgba(0, 0, 0, 0)',
                        'paper_bgcolor': 'rgba(0, 0, 0, 0)',
                        'font': {'color': '#c9d1d9'},
                        'xaxis': {'gridcolor': '#30363d', 'tickangle': 45},
                        'yaxis': {'gridcolor': '#30363d'}
                    })
                    st.plotly_chart(fig, use_container_width=True)

                # Raw data view
                st.subheader("Task Details")
                if filtered_df.empty:
                    st.warning(f"🚫 No tasks found for selected filters. Try changing the {category_column} or date range.")
                else:
                    display_cols = ['Creation Date', 'Request Id', 'Verify_user', 'Upload_Status']
                    if category_column in filtered_df.columns:
                        display_cols.append(category_column)
                    
                    # Add time if available
                    if 'Creation Time' in filtered_df.columns:
                        display_cols.append('Creation Time')
                    
                    st.dataframe(filtered_df[display_cols])

            with tab3:
                # Data quality metrics
                st.subheader("Data Quality Assessment")
                
                # Calculate completeness metrics
                completeness_data = []
                for col in ['Request Id', 'Verify_user', 'Upload_Status', category_column]:
                    if col in filtered_df.columns:
                        completeness = filtered_df[col].notna().mean() * 100
                        completeness_data.append({'Column': col, 'Completeness (%)': completeness})
                
                completeness_df = pd.DataFrame(completeness_data)
                
                if not completeness_df.empty:
                    fig = px.bar(
                        completeness_df, 
                        x='Column', 
                        y='Completeness (%)', 
                        title='Data Completeness by Column',
                        color='Completeness (%)',
                        color_continuous_scale='Viridis'
                    )
                    fig.update_layout({
                        'plot_bgcolor': 'rgba(0, 0, 0, 0)',
                        'paper_bgcolor': 'rgba(0, 0, 0, 0)',
                        'font': {'color': '#c9d1d9'},
                        'xaxis': {'gridcolor': '#30363d'},
                        'yaxis': {'gridcolor': '#30363d'}
                    })
                    st.plotly_chart(fig, use_container_width=True)
                
                # Show incomplete entries
                missing_cols = ['Request Id', 'Verify_user', 'Upload_Status']
                missing_mask = pd.DataFrame()
                for col in missing_cols:
                    if col in filtered_df.columns:
                        missing_mask[col] = filtered_df[col].isna()
                
                if not missing_mask.empty:
                    missing_data = filtered_df[missing_mask.any(axis=1)]
                    if not missing_data.empty:
                        st.subheader("Incomplete Entries")
                        st.dataframe(missing_data[['Request Id', 'Verify_user', 'Upload_Status', 'Creation Date']])
                    else:
                        st.success("No incomplete entries found!")
                
                # Overall quality score
                required_cols = ['Request Id', 'Verify_user', 'Upload_Status']
                if all(col in filtered_df.columns for col in required_cols):
                    # Check for completeness in each required column
                    valid_mask = pd.DataFrame()
                    for col in required_cols:
                        valid_mask[col] = filtered_df[col].notna()
                    
                    valid_rows = valid_mask.all(axis=1).sum()
                    total_rows = len(filtered_df)
                    quality_score = round((valid_rows / total_rows) * 100, 2) if total_rows else 0
                    st.metric("📊 Overall Data Quality Score", f"{quality_score}%")

            # Download button
            csv = filtered_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📥 Download Filtered Data as CSV",
                data=csv,
                file_name=f"{selected_client}_tasks_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv",
                mime='text/csv',
                help="Download the filtered data as a CSV file"
            )

            # Refresh timestamp
            st.caption(f"⏱️ Last refreshed: {datetime.now().strftime('%d %b %Y %I:%M %p')}")

        except Exception as e:
            st.error(f"Error processing file: {e}")
            st.info("This error might be due to issues with the Excel file format.")
    else:
        # Show welcome message when no file is uploaded
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.image("https://cdn-icons-png.flaticon.com/512/3767/3767084.png", width=150)
            st.title("Task Insights Dashboard")
            st.info("👈 Please upload an Excel file from the sidebar to get started")
            
            # Show sample data structure
            with st.expander("📋 Expected Data Format"):
                st.markdown("""
                Bye
                    Bye
                        Bye
                """)

if __name__ == "__main__":
    insight_dashboard_app()
