import io
import os
from copy import copy
from datetime import date

import pandas as pd
from openpyxl import load_workbook

class Alltrans:
    def __init__(self, template_path="Template.xlsx",
                 alltrans_sheet="All Trans", alltrans_header_row=4, mvr_sheet_name="MVR"):
        self.TEMPLATE_PATH = template_path
        self.ALLTRANS_SHEET = alltrans_sheet
        self.ALLTRANS_HEADER_ROW = alltrans_header_row
        self.MVR_PREFERRED_NAME = mvr_sheet_name

    # ---------------- Helpers ----------------
    def _replicate_sheet_across_workbooks(self, src_ws, dst_wb, dst_title):
        if dst_title in dst_wb.sheetnames:
            dst_wb.remove(dst_wb[dst_title])
        dst_ws = dst_wb.create_sheet(dst_title)
        for row in src_ws.iter_rows(values_only=False):
            for src_cell in row:
                r = src_cell.row
                c = src_cell.column
                dst_cell = dst_ws.cell(row=r, column=c, value=src_cell.value)
                try:
                    if src_cell.has_style:
                        if src_cell.font is not None:
                            dst_cell.font = copy(src_cell.font)
                        if src_cell.fill is not None:
                            dst_cell.fill = copy(src_cell.fill)
                        if src_cell.border is not None:
                            dst_cell.border = copy(src_cell.border)
                        if src_cell.alignment is not None:
                            dst_cell.alignment = copy(src_cell.alignment)
                        if src_cell.number_format is not None:
                            dst_cell.number_format = src_cell.number_format
                except Exception:
                    pass
                try:
                    if src_cell.hyperlink:
                        dst_cell._hyperlink = copy(src_cell.hyperlink)
                except Exception:
                    pass
                try:
                    if src_cell.comment:
                        dst_cell.comment = copy(src_cell.comment)
                except Exception:
                    pass
        try:
            for merged in list(src_ws.merged_cells.ranges):
                dst_ws.merge_cells(str(merged))
        except Exception:
            pass
        try:
            for col_letter, col_dim in src_ws.column_dimensions.items():
                if col_dim.width is not None:
                    dst_ws.column_dimensions[col_letter].width = col_dim.width
        except Exception:
            pass
        try:
            for r_idx, row_dim in src_ws.row_dimensions.items():
                if row_dim.height is not None:
                    dst_ws.row_dimensions[r_idx].height = row_dim.height
        except Exception:
            pass
        try:
            dst_ws.sheet_view = copy(src_ws.sheet_view)
        except Exception:
            pass
        try:
            dst_ws.freeze_panes = src_ws.freeze_panes
        except Exception:
            pass
        try:
            dst_ws.page_setup = copy(src_ws.page_setup)
            dst_ws.page_margins = copy(src_ws.page_margins)
            dst_ws.print_options = copy(src_ws.print_options)
        except Exception:
            pass
        try:
            dst_ws.sheet_properties = copy(src_ws.sheet_properties)
        except Exception:
            pass
        try:
            dst_ws.protection = copy(src_ws.protection)
        except Exception:
            pass
        try:
            if hasattr(src_ws, "tab_color"):
                dst_ws.tab_color = copy(src_ws.tab_color)
        except Exception:
            pass
        return dst_ws

    def _format_doh_for_excel(self, val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        try:
            ts = pd.to_datetime(val, errors="coerce")
            if pd.isna(ts):
                s = str(val).strip()
                return f"'{s}" if s != "" else None
            return f"{ts.strftime('%m/%d/%Y')}"
        except Exception:
            s = str(val).strip()
            return f"'{s}" if s != "" else None

    def _format_dob_mmddyyyy_str(self, val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        try:
            ts = pd.to_datetime(val, errors="coerce")
            if pd.isna(ts):
                s = str(val).strip()
                return f"'{s}" if s != "" else None
            return f"{ts.strftime('%m/%d/%Y')}"
        except Exception:
            s = str(val).strip()
            return f"'{s}" if s != "" else None

    def _normalize_tokens(self, name):
        if name is None:
            return []
        s = str(name).strip().lower()
        if s == "":
            return []
        for ch in [",", ".", "-", "_", "/"]:
            s = s.replace(ch, " ")
        return [t for t in s.split() if t]

    def _name_token_overlap(self, m_tokens, l_tokens):
        if not m_tokens or not l_tokens:
            return 0
        m_set = set(m_tokens)
        l_set = set(l_tokens)
        exact = len(m_set & l_set)
        initials = 0
        for mt in m_set:
            for lt in l_set:
                if mt and lt and mt[0] == lt[0]:
                    initials += 1
        contain = 0
        for mt in m_set:
            for lt in l_set:
                if mt in lt or lt in mt:
                    contain += 1
        return max(exact, initials, contain)

    def _name_dob_flexible_match(self, mvr_name, mvr_dob, lookup_name, lookup_dob):
        try:
            m_dt = pd.to_datetime(mvr_dob, errors="coerce")
            l_dt = pd.to_datetime(lookup_dob, errors="coerce")
            if pd.isna(m_dt) or pd.isna(l_dt):
                return False
            if m_dt.date() != l_dt.date():
                return False
        except Exception:
            return False
        m_tokens = self._normalize_tokens(mvr_name)
        l_tokens = self._normalize_tokens(lookup_name)
        if not m_tokens or not l_tokens:
            return False
        overlap = self._name_token_overlap(m_tokens, l_tokens)
        if overlap >= 2:
            return True
        if len(m_tokens) >= 2:
            return (m_tokens[0] in " ".join(l_tokens)) and (m_tokens[-1] in " ".join(l_tokens))
        return m_tokens[0] in " ".join(l_tokens)

    def _find_cdl_col(self, cols, sample_df=None):
        candidates = ["CDL Number", "CDLNumber", "CDL No", "CDL", "CDL_Number"]
        for c in candidates:
            for col in cols:
                if col.strip().lower() == c.strip().lower():
                    return col
        for col in cols:
            if "cdl" in col.strip().lower():
                return col
        if sample_df is not None:
            for col in cols:
                try:
                    sample = sample_df[col].dropna().astype(str)
                    if not sample.empty and sample.iloc[0].strip().replace(" ", "").isdigit():
                        return col
                except Exception:
                    continue
        return None

    def _find_hire_col(self, cols):
        candidates = ["Hire Date", "HireDate", "DOH", "Date of Hire", "Hire_Date"]
        for c in candidates:
            for col in cols:
                if col.strip().lower() == c.strip().lower():
                    return col
        for col in cols:
            if "hire" in col.strip().lower() or "doh" in col.strip().lower():
                return col
        return None

    def _compute_age_from_str_dob(self, s):
        if s is None:
            return None
        try:
            s2 = s[1:] if isinstance(s, str) and s.startswith("'") else s
            ts = pd.to_datetime(s2, errors="coerce")
            if pd.isna(ts):
                return None
            today = date.today()
            return today.year - ts.year - ((today.month, today.day) < (ts.month, ts.day))
        except Exception:
            return None

    # ---------------- Core run ----------------
    def run(self, main_bytes: bytes, lookup_bytes: bytes, chosen_lookup_sheet: str = None, preview_rows: int = 8):
        # load main workbook (openpyxl) for replication
        main_wb = load_workbook(io.BytesIO(main_bytes), data_only=False)
        if self.MVR_PREFERRED_NAME in main_wb.sheetnames:
            src_mvr_ws = main_wb[self.MVR_PREFERRED_NAME]
            mvr_sheet_name_used = self.MVR_PREFERRED_NAME
        else:
            src_mvr_ws = main_wb[main_wb.sheetnames[0]]
            mvr_sheet_name_used = main_wb.sheetnames[0]

        # read raw main MVR to apply skip rules
        df_raw = pd.read_excel(io.BytesIO(main_bytes), sheet_name=mvr_sheet_name_used, dtype=str, header=None)
        if len(df_raw) < 3:
            raise ValueError("Main MVR sheet too small to apply row skipping rules.")
        header_row_mvr = df_raw.iloc[1].fillna("").astype(str).tolist()
        data_rows = df_raw.drop(index=[0,2]).reset_index(drop=True)
        data_rows = data_rows.drop(index=0).reset_index(drop=True)
        df_main_clean = data_rows.copy()
        df_main_clean.columns = [h.strip() if h is not None else "" for h in header_row_mvr]
        df_main_clean.columns = [str(c).strip() for c in df_main_clean.columns]

        # lookup workbook and sheet selection
        lookup_wb = load_workbook(io.BytesIO(lookup_bytes), read_only=True, data_only=True)
        lookup_sheets = lookup_wb.sheetnames
        if chosen_lookup_sheet and chosen_lookup_sheet in lookup_sheets:
            chosen_sheet = chosen_lookup_sheet
        else:
            chosen_sheet = lookup_sheets[0]

        # preview and auto-detect header row
        df_lookup_preview = pd.read_excel(io.BytesIO(lookup_bytes), sheet_name=chosen_sheet, dtype=str, header=None, nrows=preview_rows)
        def score_row_as_header(row):
            text_vals = [str(x).strip().lower() for x in row.fillna("").astype(str).tolist()]
            score = 0
            tokens = ("cdl","hire","doh","driver","name","dob","birth")
            for t in text_vals:
                for ex in tokens:
                    if ex in t:
                        score += 1
            return score
        header_scores = [score_row_as_header(df_lookup_preview.iloc[i]) for i in range(len(df_lookup_preview))]
        best_idx = int(pd.Series(header_scores).idxmax())
        best_score = header_scores[best_idx]
        auto_header_row = best_idx if best_score >= 2 else 0
        lookup_df = pd.read_excel(io.BytesIO(lookup_bytes), sheet_name=chosen_sheet, dtype=str, header=auto_header_row)
        lookup_df.columns = [str(c).strip() for c in lookup_df.columns]

        lookup_cols = list(lookup_df.columns)
        cdl_col_lookup = self._find_cdl_col(lookup_cols, sample_df=lookup_df)
        hire_col_lookup = self._find_hire_col(lookup_cols)
        if cdl_col_lookup is None or hire_col_lookup is None:
            raise ValueError("Could not detect CDL or Hire columns in lookup automatically.")

        # normalize lookup keys and detect duplicates
        lookup_df["_cdl_key"] = lookup_df[cdl_col_lookup].astype(str).str.strip()
        dup_mask = lookup_df["_cdl_key"].duplicated(keep=False)
        dupe_rows = lookup_df[dup_mask].copy()
        dupe_summary = dupe_rows["_cdl_key"].value_counts().to_dict() if not dupe_rows.empty else {}
        lookup_small = lookup_df.drop_duplicates("_cdl_key", keep="first").copy()
        lookup_rows = lookup_df.to_dict(orient="records")

        # normalize lookup rows: name, dob (MM/DD/YYYY string), keep hire raw
        for r in lookup_rows:
            r["_cdl_key_norm"] = str(r.get("_cdl_key", "")).strip()
            r["_lookup_name"] = ""
            r["_lookup_dob"] = None
            for c in lookup_df.columns:
                if "name" in str(c).lower() and (r.get(c) is not None and str(r.get(c)).strip() != ""):
                    r["_lookup_name"] = r.get(c)
                    break
            for c in lookup_df.columns:
                if any(k in str(c).lower() for k in ("dob","date of birth","birth")):
                    try:
                        dt = pd.to_datetime(r.get(c), errors="coerce")
                        r["_lookup_dob"] = dt.strftime("%m/%d/%Y") if not pd.isna(dt) else None
                    except Exception:
                        r["_lookup_dob"] = None
                    break
            r["_hire_raw"] = r.get(hire_col_lookup)

        # build direct CDL map
        lookup_map = pd.Series(lookup_small[hire_col_lookup].values, index=lookup_small["_cdl_key"]).to_dict()

        # prepare main keys & order
        cdl_col_main = self._find_cdl_col(list(df_main_clean.columns), sample_df=df_main_clean)
        if cdl_col_main is None:
            raise ValueError("Could not detect CDL Number column in main MVR cleaned sheet.")
        df_main_clean["_cdl_key"] = df_main_clean[cdl_col_main].astype(str).str.strip()
        order_keys = df_main_clean["_cdl_key"].dropna().astype(str).tolist()
        seen = set()
        ordered_unique_keys = []
        for k in order_keys:
            if k not in seen:
                seen.add(k)
                ordered_unique_keys.append(k)

        # pick main columns
        cols = list(df_main_clean.columns)
        def pick(col_candidates):
            for cand in col_candidates:
                for c in cols:
                    if c.strip().lower() == cand.strip().lower():
                        return c
            for cand in col_candidates:
                for c in cols:
                    if cand.strip().lower() in c.strip().lower():
                        return c
            return None

        driver_col = pick(["Driver Full Name","DriverFullName","Name of Driver","Driver"])
        dob_col = pick(["Driver Date of Birth","Date of birth","DOB"])
        cdl_type_col = pick(["CDL Type","CDLType","Lic Class","License Class"])
        lic_state_col = pick(["License State","LicenseState","State"])
        pass_end_col = pick(["License Endorsement - Passenger","Passenger Endorsement","Passenger Endt"])
        school_end_col = pick(["License Endorsement - School Bus","School Bus Endorsement","School Bus Endt"])
        viol_desc_col = pick(["Violation Description","ViolationDescription","Violation Desc","Description"])
        viol_cat_col = pick(["Violation Category","Violation Type","Category"])
        lic_status_col = pick(["License Status","Lic Status","Status"])

        # aggregate per CDL and normalize MVR DOB to MM/DD/YYYY with leading apostrophe
        grouped = df_main_clean.groupby("_cdl_key", sort=False)
        records = []
        for key in ordered_unique_keys:
            group = grouped.get_group(key) if key in grouped.groups else pd.DataFrame(columns=df_main_clean.columns)
            rec = {}
            rec["_cdl_key"] = key
            rec["Driver Full Name"] = group[driver_col].dropna().astype(str).iloc[0] if (driver_col in group and not group[driver_col].dropna().empty) else None
            raw_dob = group[dob_col].dropna().iloc[0] if (dob_col in group and not group[dob_col].dropna().empty) else None
            try:
                dt = pd.to_datetime(raw_dob, errors="coerce")
                dob_norm = dt.strftime("%m/%d/%Y") if not pd.isna(dt) else (str(raw_dob).strip() if raw_dob is not None else None)
                rec["Driver Date of Birth"] = f"{dob_norm}" if dob_norm else None
            except Exception:
                rec["Driver Date of Birth"] = f"{str(raw_dob).strip()}" if raw_dob not in (None, "") else None
            rec["CDL Number"] = key
            rec["CDL Type"] = group[cdl_type_col].dropna().iloc[0] if (cdl_type_col in group and not group[cdl_type_col].dropna().empty) else None
            rec["Lic State"] = group[lic_state_col].dropna().iloc[0] if (lic_state_col in group and not group[lic_state_col].dropna().empty) else None
            rec["Passenger Endt"] = group[pass_end_col].dropna().iloc[0] if (pass_end_col in group and not group[pass_end_col].dropna().empty) else None
            rec["School Bus Endt"] = group[school_end_col].dropna().iloc[0] if (school_end_col in group and not group[school_end_col].dropna().empty) else None
            rec["LIC Status"] = group[lic_status_col].dropna().astype(str).iloc[0] if (lic_status_col in group and not group[lic_status_col].dropna().empty) else None
            notes = []
            if viol_desc_col in group:
                notes = [str(x).strip() for x in group[viol_desc_col].dropna().astype(str).unique()]
            rec["Notes"] = " ; ".join([n for n in notes if n and n.lower() != "nan"]) if notes else None
            cat_series = group[viol_cat_col].astype(str) if (viol_cat_col and viol_cat_col in group) else pd.Series([""] * len(group))
            desc_series = group[viol_desc_col].astype(str) if (viol_desc_col and viol_desc_col in group) else pd.Series([""] * len(group))
            rec["# Accidents"] = int(cat_series.str.contains("accident", case=False, na=False).sum()) if not cat_series.empty else int(desc_series.str.contains("accident", case=False, na=False).sum())
            rec["# Minor Violations"] = int(cat_series.str.contains("minor", case=False, na=False).sum()) if not cat_series.empty else int(desc_series.str.contains("minor", case=False, na=False).sum())
            rec["# MAJOR Violations"] = int(cat_series.str.contains("|".join(["major","dui","dwi","felony","suspension"]), case=False, na=False).sum()) if not cat_series.empty else int(desc_series.str.contains("|".join(["major","dui","dwi","felony","suspension"]), case=False, na=False).sum())
            rec["YES"] = "X" if str(rec.get("LIC Status","")).strip().upper() == "VALID" else None
            rec["NO"] = None
            records.append(rec)
        df_records = pd.DataFrame(records)

        # matching: CDL primary, Name+DOB fallback; keep track of matched lookup rows
        fallback_matches = []
        unmatched_mask_indices = []
        matched_lookup_indices = set()
        for idx, rec in df_records.iterrows():
            cdl_key = str(rec.get("_cdl_key", "")).strip()
            doh_raw = None
            matched_by = None
            if cdl_key and cdl_key in lookup_map:
                doh_raw = lookup_map.get(cdl_key)
                matched_by = "CDL"
                for i, lr in enumerate(lookup_rows):
                    if str(lr.get("_cdl_key_norm","")).strip() == cdl_key:
                        matched_lookup_indices.add(i)
            else:
                m_name = rec.get("Driver Full Name")
                m_dob_str = rec.get("Driver Date of Birth")
                m_dob_compare = m_dob_str[1:] if (isinstance(m_dob_str, str) and m_dob_str.startswith("'")) else m_dob_str
                found = None
                for i, lr in enumerate(lookup_rows):
                    if i in matched_lookup_indices:
                        continue
                    lk_name = lr.get("_lookup_name") or lr.get(next((c for c in lookup_df.columns if "driver" in str(c).lower()), None), "")
                    lk_dob_str = lr.get("_lookup_dob")
                    if self._name_dob_flexible_match(m_name, m_dob_compare, lk_name, lk_dob_str):
                        found = (i, lr)
                        break
                if found:
                    i, lr = found
                    doh_raw = lr.get(hire_col_lookup)
                    matched_by = "Name+DOB"
                    fallback_matches.append({
                        "MVR_CDL": cdl_key or None,
                        "Driver": rec.get("Driver Full Name"),
                        "MVR_DOB": m_dob_compare,
                        "Matched Lookup CDL": lr.get("_cdl_key_norm"),
                        "Matched Lookup Name": lr.get("_lookup_name"),
                        "Matched Hire Raw": doh_raw
                    })
                    matched_lookup_indices.add(i)
                else:
                    matched_by = None
                    unmatched_mask_indices.append(idx)
            df_records.at[idx, "DOH_raw"] = doh_raw
            df_records.at[idx, "DOH"] = self._format_doh_for_excel(doh_raw) if doh_raw is not None else None
            df_records.at[idx, "MatchedBy"] = matched_by

        # append lookup-only rows (those not matched)
        appended_lookup_rows = []
        for i, lr in enumerate(lookup_rows):
            if i in matched_lookup_indices:
                continue
            append_rec = {}
            append_rec["_cdl_key"] = lr.get("_cdl_key_norm") or None
            name_val = lr.get("_lookup_name") or lr.get(next((c for c in lookup_df.columns if "name" in str(c).lower()), None), None)
            append_rec["Driver Full Name"] = name_val
            dob_val = lr.get("_lookup_dob") or None
            append_rec["Driver Date of Birth"] = f"{dob_val}" if dob_val not in (None, "") else None
            append_rec["CDL Number"] = lr.get("_cdl_key_norm")
            append_rec["CDL Type"] = None
            append_rec["Lic State"] = None
            append_rec["Passenger Endt"] = None
            append_rec["School Bus Endt"] = None
            append_rec["LIC Status"] = None
            append_rec["Notes"] = "Missing MVR"
            append_rec["# Accidents"] = 0
            append_rec["# Minor Violations"] = 0
            append_rec["# MAJOR Violations"] = 0
            append_rec["YES"] = None
            append_rec["NO"] = None
            doh_raw = lr.get(hire_col_lookup)
            append_rec["DOH_raw"] = doh_raw
            append_rec["DOH"] = self._format_doh_for_excel(doh_raw) if doh_raw is not None else None
            append_rec["MatchedBy"] = "LookupOnly"
            appended_lookup_rows.append(append_rec)
        if appended_lookup_rows:
            df_appends = pd.DataFrame(appended_lookup_rows)
            df_records = pd.concat([df_records, df_appends], ignore_index=True)

        # compute Age from DOB (strip apostrophe)
        df_records["Age"] = df_records["Driver Date of Birth"].apply(self._compute_age_from_str_dob)

        # build issues
        issues = []
        missing_cdl_mask = df_records["_cdl_key"].isna() | (df_records["_cdl_key"].astype(str).str.strip() == "")
        if missing_cdl_mask.any():
            for _, r in df_records[missing_cdl_mask].iterrows():
                issues.append({"Type": "Missing CDL in MVR", "CDL": r.get("_cdl_key"), "Driver": r.get("Driver Full Name"), "Note": "Blank or missing CDL in MVR row"})
        missing_doh_mask = df_records["DOH_raw"].isna() | (df_records["DOH_raw"].astype(str).str.strip() == "")
        if missing_doh_mask.any():
            for _, r in df_records[missing_doh_mask].iterrows():
                issues.append({"Type": "Missing DOH", "CDL": r.get("_cdl_key"), "Driver": r.get("Driver Full Name"), "Note": "No hire date found in lookup by CDL or Name+DOB"})
        if dupe_summary:
            for k, cnt in dupe_summary.items():
                issues.append({"Type": "Duplicate Lookup CDL", "CDL": k, "Driver": None, "Note": f"{cnt} rows found in lookup for this CDL; first was used"})
        for fm in fallback_matches:
            issues.append({"Type": "Fallback match Name+DOB", "CDL": fm.get("MVR_CDL"), "Driver": fm.get("Driver"), "Note": f"Matched to lookup CDL {fm.get('Matched Lookup CDL')}; hire raw: {fm.get('Matched Hire Raw')}"})
        for ar in appended_lookup_rows:
            issues.append({"Type": "Appended Lookup-only driver", "CDL": ar.get("_cdl_key"), "Driver": ar.get("Driver Full Name"), "Note": "Appended from lookup with Notes='Missing MVR'"})
        lookup_only_keys = set([str(k) for k in lookup_small["_cdl_key"].astype(str)]) - set([str(k) for k in df_main_clean["_cdl_key"].astype(str)])
        if lookup_only_keys:
            for k in sorted(list(lookup_only_keys))[:200]:
                issues.append({"Type": "Lookup CDL not in MVR (informational)", "CDL": k, "Driver": None, "Note": "Exists in lookup but not in original MVR"})

        # write to template
        if not os.path.exists(self.TEMPLATE_PATH):
            raise FileNotFoundError(f"Template not found at {self.TEMPLATE_PATH}")
        wb = load_workbook(self.TEMPLATE_PATH)
        if self.ALLTRANS_SHEET not in wb.sheetnames:
            raise ValueError(f"Template must contain sheet '{self.ALLTRANS_SHEET}'")
        ws = wb[self.ALLTRANS_SHEET]
        header_cells = list(ws[self.ALLTRANS_HEADER_ROW])
        template_headers = [c.value for c in header_cells]
        data_start_row = self.ALLTRANS_HEADER_ROW + 1

        for idx, rec in df_records.reset_index(drop=True).iterrows():
            write_row = data_start_row + idx
            for col_idx, hdr in enumerate(template_headers, start=1):
                if hdr is None:
                    continue
                h = str(hdr).strip().lower()
                val = None
                if h in ("name of driver","driver full name","name","driver"):
                    val = rec.get("Driver Full Name")
                elif h in ("dob","date of birth"):
                    val = rec.get("Driver Date of Birth")
                elif h == "age":
                    val = rec.get("Age")
                elif h in ("lic state","license state","state"):
                    val = rec.get("Lic State")
                elif h in ("lic class","license class","cdl type","class"):
                    val = rec.get("CDL Type")
                elif "passenger" in h:
                    val = rec.get("Passenger Endt")
                elif "school" in h:
                    val = rec.get("School Bus Endt")
                elif "accident" in h:
                    val = rec.get("# Accidents")
                elif "minor" in h:
                    val = rec.get("# Minor Violations")
                elif "major" in h:
                    val = rec.get("# MAJOR Violations")
                elif h.upper() == "YES":
                    val = rec.get("YES")
                elif h.upper() == "NO":
                    val = None
                elif h in ("notes",):
                    existing_notes = rec.get("Notes") or ""
                    if (rec.get("MatchedBy") == "LookupOnly") or (pd.isna(rec.get("DOH_raw")) or rec.get("DOH_raw") is None or str(rec.get("DOH_raw")).strip() == ""):
                        if existing_notes:
                            val = f"{existing_notes} ; Missing MVR"
                        else:
                            val = "Missing MVR"
                    else:
                        val = existing_notes if existing_notes else None
                elif h in ("doh","date of hire","hire date"):
                    val = rec.get("DOH")
                elif "af claims" in h:
                    val = None
                elif "lic status" in h or "license status" in h:
                    val = rec.get("LIC Status")
                else:
                    if hdr in df_records.columns:
                        val = rec.get(hdr)
                try:
                    ws.cell(row=write_row, column=col_idx, value=val)
                except Exception:
                    pass

        # replicate MVR sheet exactly
        self._replicate_sheet_across_workbooks(src_mvr_ws, wb, "MVR")

        # write Issues sheet
        if "Issues" in wb.sheetnames:
            wb.remove(wb["Issues"])
        issues_ws = wb.create_sheet("Issues")
        summary = {
            "Total drivers written": len(df_records),
            "Drivers missing CDL": int(missing_cdl_mask.sum()) if 'missing_cdl_mask' in locals() else 0,
            "Drivers missing DOH after matching": int(missing_doh_mask.sum()) if 'missing_doh_mask' in locals() else 0,
            "Duplicate lookup CDLs": len(dupe_summary)
        }
        row = 1
        for k, v in summary.items():
            issues_ws.cell(row=row, column=1, value=k)
            issues_ws.cell(row=row, column=2, value=v)
            row += 1
        row += 1
        headers = ["Type", "CDL", "Driver", "Note"]
        for c_idx, h in enumerate(headers, start=1):
            issues_ws.cell(row=row, column=c_idx, value=h)
        row += 1
        for it in issues:
            issues_ws.cell(row=row, column=1, value=it.get("Type"))
            issues_ws.cell(row=row, column=2, value=it.get("CDL"))
            issues_ws.cell(row=row, column=3, value=it.get("Driver"))
            issues_ws.cell(row=row, column=4, value=it.get("Note"))
            row += 1

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        return out
